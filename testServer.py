import threading
import socket
import time
import requests
import schedule
from bs4 import BeautifulSoup
import openpyxl
from concurrent.futures import ThreadPoolExecutor

EXCEL_FILE = "price_data.xlsx"

class ItemThread(threading.Thread):
    def __init__(self, client_socket, product_name, desired_price, product_link, daily_average_time):
        super(ItemThread, self).__init__()
        self.client_socket = client_socket
        self.product_name = product_name
        self.desired_price = desired_price
        self.product_link = product_link
        self.daily_average_time = daily_average_time
        self.running = True
        self.crawled_price = None
        self.hourly_prices = []
        self.price_data_list = []

    def crawling_test(self):
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
                "Accept-Language": "ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3"
            }
            res = requests.get(self.product_link, headers=headers)
            res.raise_for_status()

            soup = BeautifulSoup(res.text, "html.parser")
            price_element = soup.select_one(".total-price")

            if price_element:
                self.crawled_price = int(price_element.get_text(strip=True).replace(",", "").replace("원", ""))
                self.client_socket.send(str(self.crawled_price).encode())
                current_time = time.strftime("%H:%M")
                self.hourly_prices.append(self.crawled_price)
                return self.crawled_price
            else:
                self.crawled_price = None
                return self.crawled_price
        except Exception as e:
            print(f"Exception occurred during crawling: {e}")
            self.crawled_price = None
            return self.crawled_price

    def calculate_daily_average(self):
        try:
            total_price = sum(self.hourly_prices)
            num_of_data_points = len(self.hourly_prices)
            if num_of_data_points == 0:
                print("No data points found for the day.")
                return
            daily_average = int(total_price / num_of_data_points)
            print(f"Daily Average Price: {daily_average}")
            self.price_data_list.append(("일평균", daily_average))
            self.save_to_excel(self.price_data_list)
            self.hourly_prices.clear()
            self.price_data_list.clear()
        except Exception as e:
            print(f"Exception occurred during daily average calculation: {e}")


    def save_to_excel(self, price_data):
        current_date = time.strftime("%m/%d")
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        
        if self.product_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(self.product_name)
        else:
            sheet = workbook[self.product_name]
        
        if not sheet['A1'].value:
            sheet.append(["날짜"] + [f"{hour:02d}:00" for hour in range(24)] + ["일평균가"])
        
        row_data = [current_date] + [price for _, price in price_data]
        sheet.append(row_data)
        
        workbook.save(EXCEL_FILE)
        workbook.close()

    def showCurrentPrice(self):
        self.crawlingTest()
        print(f"Current price: {self.crawled_price}")
        self.client_socket.send(str(self.crawled_price).encode())
      
    def run(self):
        schedule.every().day.at(self.daily_average_time).do(self.calculate_daily_average)
        for hour in range(24):
            time_str = f"{hour:02d}:00"
            schedule.every().day.at(time_str).do(self.crawling_test).tag(time_str)

class PriceServer:
    def __init__(self):
        self.init_excel()
        self.init_server()

    def init_excel(self):
        try:
            openpyxl.load_workbook(EXCEL_FILE).close()
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            workbook.save(EXCEL_FILE)
            workbook.close()

    def init_server(self):
        self.server_host = 'localhost'
        self.server_port = 12345
        self.server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        self.server_socket.bind((self.server_host, self.server_port))
        self.server_socket.listen(5)
        print(f"Server listening on {self.server_host}:{self.server_port}")
        self.item_daily_average_time = "23:50"
        self.thread_pool = ThreadPoolExecutor(max_workers=5)

    def handle_client(self, client_socket, address):
        while True:
            data = client_socket.recv(1024).decode()
            if not data:
                break
            product_name, desired_price, product_link = data.split(',')
            item_thread = ItemThread(client_socket, product_name, desired_price, product_link, self.item_daily_average_time)
            item_thread.start()
        client_socket.close()

    def start(self):
        while True:
            client_socket, address = self.server_socket.accept()
            print(f"Accepted connection from {address}")
            data = client_socket.recv(1024).decode()
            if not data:
                client_socket.close()
                continue
            product_name, desired_price, product_link = data.split(',')
            item_thread = ItemThread(client_socket, product_name, desired_price, product_link, self.item_daily_average_time)
            self.thread_pool.submit(item_thread.start)

if __name__ == "__main__":
    price_server = PriceServer()
    price_server.start()
