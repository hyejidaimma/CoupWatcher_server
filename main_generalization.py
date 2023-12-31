import threading
import socket
import time
import requests
import schedule as schedule
from bs4 import BeautifulSoup
import openpyxl
from concurrent.futures import ThreadPoolExecutor

# 엑셀 파일에 대한 정보를 전역 변수로 선언(일반화)
EXCEL_FILE = "price_data.xlsx"
SHEET_NAME = "PriceData"

# 열의 종류를 미리 설정
COLUMN_HEADERS = ["날짜", "00:00", "01:00", "02:00", "03:00", "04:00", "05:00", "06:00",
                  "07:00", "08:00", "09:00", "10:00", "11:00", "12:00", "13:00", "14:00",
                  "15:00", "16:00", "17:00", "18:00", "19:00", "20:00", "21:00", "22:00",
                  "23:00", "일평균가"]


class ItemThread(threading.Thread):
    def __init__(self, client_socket, product_name, desired_price, product_link, daily_average_time):
        super(ItemThread, self).__init__()
        self.client_socket = client_socket
        self.product_name = product_name
        self.desired_price = desired_price
        self.product_link = product_link
        self.daily_average_time = daily_average_time
        self.running = True
        self.crawled_price = None
        self.crawled_count = 0
        self.average_price = 0
        self.hourly_prices = []     #일평균가를 계산하기 위한 시간대별 가격 리스트
        self.price_data_list = []   # 날짜, 시간대별 가격, 일평균 가격을 저장하는 리스트

    def crawlingTest(self):
        try:
            #쿠팡 크롤링위해 꼭 필요(로봇 아님 인증)
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
                "Accept-Language": "ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3"
            }

            res = requests.get(self.product_link, headers=headers)
            res.raise_for_status()

            soup = BeautifulSoup(res.text, "html.parser")

            # 가격 정보를 추출
            price_element = soup.select_one(".total-price")

            if price_element:
                # 문자열을 정수로 변환
                self.crawled_price = int(price_element.get_text(strip=True).replace(",", "").replace("원", ""))
                # (파이썬)클라이언트 전송 테스트용
                self.client_socket.send(str(self.crawled_price).encode())

                # 날짜와 시간대별 가격을 price_data_list에 추가
                current_time = time.strftime("%H:%M")
                self.price_data_list.append((current_time, self.crawled_price))
                return self.crawled_price
            else:
                self.crawled_price = None  # 예외 발생 시 가격 정보를 None으로 설정
                return self.crawled_price
        except Exception as e:
            print(f"Exception occurred during crawling: {e}")
            self.crawled_price = None  # 예외 발생 시 가격 정보를 None으로 설정
            return self.crawled_price

    def calculate_daily_average(self):
        try:
            total_price = sum(self.hourly_prices)
            num_of_data_points = len(self.hourly_prices)

            if num_of_data_points == 0:
                print("No data points found for the day.")
                return

            # Daily average calculation
            daily_average =int(total_price / num_of_data_points)
            print(f"Daily Average Price: {daily_average}")

            # 일평균 가격을 price_data_list에 추가
            # current_date = time.strftime("%m/%d")
            self.price_data_list.append(("일평균", daily_average))

            # 엑셀 파일에 데이터 저장
            self.save_to_excel(self.price_data_list)

            # Clear the list for the next day
            self.hourly_prices.clear()
            self.price_data_list.clear()
        except Exception as e:
            print(f"Exception occurred during daily average calculation: {e}")

    def save_to_excel(self, price_data):
        current_date = time.strftime("%m/%d")

        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook[SHEET_NAME]

        # 첫 번째 행에 열의 종류를 추가
        if not sheet['A1'].value:
            sheet.append(COLUMN_HEADERS)

        # price_data_list를 엑셀 파일에 추가
        row_data = [current_date] + [price for _, price in price_data]
        sheet.append(row_data)

        workbook.save(EXCEL_FILE)
        workbook.close()

    def showCurrentPrice(self):
        self.crawlingTest()
        print(f"Current price: {self.crawled_price}")
        self.client_socket.send(str(self.crawled_price).encode())

    def run(self):
        # 00:00부터 23:00까지 매 시 정각마다 크롤링을 수행하고 가격을 hourly_prices에 저장
        for hour in range(24):
            time_str = f"{hour:02d}:00"
            schedule.every().day.at(time_str).do(self.crawlingTest).tag(time_str)  # Add tag for each job

        # 매일 23:50에 calculate_daily_average() 함수를 실행하여 일평균 가격을 계산
        schedule.every().day.at("23:50").do(self.calculate_daily_average)

class PriceServer:
    def __init__(self):
        try:
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            workbook.close()
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            workbook.save(EXCEL_FILE)
            workbook.close()

        # 서버 설정
        self.server_host = 'localhost'
        self.server_port = 12345
        self.server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        self.server_socket.bind((self.server_host, self.server_port))
        self.server_socket.listen(5)
        print(f"Server listening on {self.server_host}:{self.server_port}")

        # 아이템별 일평균 계산 시간 설정
        self.item_daily_average_time = "23:50"  # 매일 23:50에 일평균 계산

        # 스레드풀 생성 (최대 10개의 스레드)
        self.thread_pool = ThreadPoolExecutor(max_workers=5)

    def handle_client(self, client_socket, address):
        while True:
            data = client_socket.recv(1024).decode()
            if not data:
                break

            product_name, desired_price, product_link = data.split(',')
            item_thread = ItemThread(client_socket, product_name, desired_price, product_link,
                                     self.item_daily_average_time)
            item_thread.start()

        client_socket.close()

    def start(self):
        while True:
            client_socket, address = self.server_socket.accept()
            print(f"Accepted connection from {address}")

            data = client_socket.recv(1024).decode()
            if not data:
                client_socket.close()
                continue

            product_name, desired_price, product_link = data.split(',')
            item_thread = ItemThread(client_socket, product_name, desired_price, product_link,
                                     self.item_daily_average_time)
            # 스레드풀에 스레드 추가 및 실행
            self.thread_pool.submit(item_thread.start)


if __name__ == "__main__":
    price_server = PriceServer()
    price_server.start()
